from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.incident import Category, Incident, Status
from app.services.incidents import list_incidents

router = APIRouter(prefix="/api/export", tags=["export"])

HEADERS = [
    "ID", "Typ", "Data zdarzenia", "Oddział", "Lokalizacja", "Kategoria",
    "Opis", "Ciężkość", "Działania podjęte", "Opis działań",
    "Propozycje zapobiegawcze", "Wiek pacjenta", "Płeć", "Zgłaszający",
    "Rola", "Anonimowe", "Status", "Data zgłoszenia",
]

SEVERITY_MAP = {0: "Brak szkody", 1: "Minimalna", 2: "Umiarkowana", 3: "Poważna", 4: "Krytyczna/zgon"}
EVENT_TYPE_MAP = {"HARMFUL": "ZN", "NO_HARM": "ZN-0", "NEAR_MISS": "NZN", "ZN": "ZN", "ZN-0": "ZN-0", "NZN": "NZN"}
CATEGORY_MAP = {
    "A": "Procedury kliniczne", "B": "Farmakoterapia", "C": "Zakażenia",
    "D": "Sprzęt medyczny", "E": "Upadki", "F": "Odleżyny",
    "G": "Krew", "H": "Opieka", "I": "Dokumentacja",
    "J": "Samouszkodzenie", "K": "Infrastruktura", "L": "Organizacja",
    "M": "Prawa pacjenta", "X": "Inne",
}
STATUS_MAP = {
    "new": "Nowe", "in_triage": "W triage", "rejected": "Odrzucone",
    "in_analysis": "W analizie", "escalated_rca": "Eskalowane RCA",
    "action_plan": "Plan działań", "implementing": "Wdrażanie", "closed": "Zamknięte",
}


def _val(enum_or_str):
    return enum_or_str.value if hasattr(enum_or_str, "value") else str(enum_or_str)


@router.get("/incidents.xlsx")
def export_incidents(
    status: Status | None = None,
    category: Category | None = None,
    db: Session = Depends(get_db),
):
    items, _ = list_incidents(db, status=status, category=category, skip=0, limit=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Zdarzenia niepożądane"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, inc in enumerate(items, 2):
        et = _val(inc.event_type)
        cat = _val(inc.category)
        sev = _val(inc.severity)
        st = _val(inc.status)

        values = [
            f"ZN-{str(inc.id).zfill(4)}",
            EVENT_TYPE_MAP.get(et, et),
            inc.event_date.strftime("%Y-%m-%d %H:%M"),
            inc.department,
            inc.location or "",
            f"{cat}. {CATEGORY_MAP.get(cat, cat)}",
            inc.description,
            SEVERITY_MAP.get(int(sev) if str(sev).isdigit() else sev, str(sev)),
            "Tak" if inc.immediate_actions_taken else "Nie",
            inc.immediate_actions_desc or "",
            inc.preventive_suggestions or "",
            inc.patient_age if inc.patient_age is not None else "",
            inc.patient_sex or "",
            inc.reporter_name or "",
            inc.reporter_role or "",
            "Tak" if inc.reporter_anonymous else "Nie",
            STATUS_MAP.get(st, st),
            inc.created_at.strftime("%Y-%m-%d %H:%M"),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["Q"].width = 16
    ws.column_dimensions["R"].width = 18

    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=zdarzenia_niepozadane.xlsx"},
    )
